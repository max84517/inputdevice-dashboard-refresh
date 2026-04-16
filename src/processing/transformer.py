"""
Processing module: read a supplier Excel sheet, transform wide → long,
compute derived columns, return a clean DataFrame.
"""
import os
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd

# Months mapping
MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "may": 5, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Value column patterns (case-insensitive prefix match)
VALUE_PATTERNS = [
    re.compile(r"^table price\s+(\w+)$", re.I),
    re.compile(r"^unit rebate\s+(\w+)$", re.I),
    re.compile(r"^(\w+)\s+q'?ty$", re.I),
    re.compile(r"^(\w+)\s+rebate amount$", re.I),
]

VALUE_TYPE_MAP = {
    "table price": "Table Price",
    "unit rebate": "Unit Rebate",
    "qty": "Q'ty",
    "rebate amount": "Rebate Amount",
}


def _classify_col(col_name: str):
    """Return (value_type, month_abbr) or None if not a value column."""
    col = str(col_name).strip()
    # Table Price <month>
    m = re.match(r"^table price\s+(\w+)$", col, re.I)
    if m:
        mon = m.group(1).lower()
        if mon in MONTH_MAP:
            return ("Table Price", mon)
    # Unit Rebate <month>
    m = re.match(r"^unit rebate\s+(\w+)$", col, re.I)
    if m:
        mon = m.group(1).lower()
        if mon in MONTH_MAP:
            return ("Unit Rebate", mon)
    # <month> Q'ty  or  <month> Qty
    m = re.match(r"^(\w+)\s+q'?ty$", col, re.I)
    if m:
        mon = m.group(1).lower()
        if mon in MONTH_MAP:
            return ("Q'ty", mon)
    # <month> Rebate Amount
    m = re.match(r"^(\w+)\s+rebate amount$", col, re.I)
    if m:
        mon = m.group(1).lower()
        if mon in MONTH_MAP:
            return ("Rebate Amount", mon)
    return None


def get_fy_sheets(file_path: str) -> list[str]:
    """Return sheet names matching FY + digits exactly (e.g. FY25, FY26)."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if re.match(r"^FY\d+$", s, re.I)]
    wb.close()
    return sheets


def read_sheet(file_path: str, sheet_name: str) -> pd.DataFrame:
    """
    Read a supplier sheet with openpyxl (data_only=True to get formula results).
    Header is row 2 (0-indexed: row index 1).
    Returns raw DataFrame.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return pd.DataFrame()

    raw_headers = [str(h).strip() if h is not None else "" for h in rows[1]]
    # Deduplicate column names — duplicate/blank headers cause reindex errors
    seen: dict[str, int] = {}
    headers = []
    for h in raw_headers:
        key = h if h else "__col"
        count = seen.get(key, 0)
        headers.append(h if count == 0 else f"{key}__{count}")
        seen[key] = count + 1
    data = rows[2:]
    wb.close()
    return pd.DataFrame(data, columns=headers)


def process_supplier_sheet(file_path: str, sheet_name: str, supplier_name: str) -> pd.DataFrame:
    """
    Full pipeline for one supplier sheet.
    Uses vectorized pd.melt instead of iterrows for speed.
    Returns long-format DataFrame with all derived columns.
    """
    df = read_sheet(file_path, sheet_name)
    if df.empty:
        return df

    # Identify value columns vs feature columns
    value_col_meta: dict[str, tuple[str, str]] = {}  # col_name -> (value_type, month_abbr)
    feature_cols: list[str] = []

    for col in df.columns:
        result = _classify_col(col)
        if result:
            value_col_meta[col] = result
        else:
            feature_cols.append(col)

    # Drop rows where Platforms is blank
    platforms_col = next((c for c in feature_cols if str(c).strip().lower() == "platforms"), None)
    if platforms_col is not None:
        df = df[df[platforms_col].notna() & (df[platforms_col].astype(str).str.strip() != "")]

    if df.empty:
        return df

    df = df.reset_index(drop=True)

    # Forward-fill & strip feature columns (merged-cell handling)
    for col in feature_cols:
        df[col] = df[col].ffill()
        df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

    # Drop feature columns whose name contains "FY" (e.g. "FY24", "FY Year")
    feature_cols = [c for c in feature_cols if "FY" not in str(col).upper() or col == col]
    feature_cols = [c for c in feature_cols if "FY" not in str(c).upper()]

    # ── Vectorized wide → long using pd.melt ─────────────────────────────────
    # Build per-type rename maps: {original_col: month_abbr}
    type_cols: dict[str, dict[str, str]] = {
        "Table Price": {},
        "Unit Rebate": {},
        "Q'ty": {},
        "Rebate Amount": {},
    }
    for col, (vtype, mon) in value_col_meta.items():
        type_cols[vtype][col] = mon

    # Melt each value type separately, then merge on (feature_cols + month)
    melted_parts: list[pd.DataFrame] = []
    out_names = {
        "Table Price": "HP Cost",
        "Unit Rebate": "Unit Rebate",
        "Q'ty": "Q'ty",
        "Rebate Amount": "Rebate Amount",
    }

    for vtype, col_map in type_cols.items():
        if not col_map:
            continue
        cols_to_melt = list(col_map.keys())
        # Rename to month abbr before melting so 'value' column carries month
        rename_map = {c: col_map[c] for c in cols_to_melt}
        sub = df[feature_cols + cols_to_melt].copy()
        sub.rename(columns=rename_map, inplace=True)
        # After rename, melt month columns
        month_cols = list(col_map.values())
        melted = sub.melt(id_vars=feature_cols, value_vars=month_cols,
                          var_name="__month", value_name=out_names[vtype])
        melted[out_names[vtype]] = pd.to_numeric(melted[out_names[vtype]], errors="coerce")
        melted_parts.append(melted)

    if not melted_parts:
        return pd.DataFrame()

    # Merge all value types on feature_cols + __month
    long_df = melted_parts[0]
    for part in melted_parts[1:]:
        long_df = long_df.merge(part, on=feature_cols + ["__month"], how="outer")

    # Fill blank value columns with 0
    for _vc in ["HP Cost", "Unit Rebate", "Q'ty", "Rebate Amount"]:
        if _vc in long_df.columns:
            long_df[_vc] = long_df[_vc].fillna(0)

    # Derive columns
    long_df["ODM Cost"] = long_df["HP Cost"] + long_df["Unit Rebate"]
    long_df["Spending Amount"] = long_df["ODM Cost"] * long_df["Q'ty"]
    long_df["Actual Spending"] = long_df["HP Cost"] * long_df["Q'ty"]

    # Month number
    long_df["Month"] = long_df["__month"].map(
        lambda m: MONTH_MAP.get(m.lower() if isinstance(m, str) else "", None)
    )
    long_df.drop(columns=["__month"], inplace=True)

    # FY year from sheet name e.g. FY26 -> 2026
    fy_num = int(re.search(r"\d+", sheet_name).group()) + 2000

    long_df["Year"] = long_df["Month"].map(
        lambda m: (fy_num - 1) if m in (11, 12) else fy_num
    )

    def _quarter(m):
        if m in (11, 12, 1): return "Q1"
        elif m in (2, 3, 4):  return "Q2"
        elif m in (5, 6, 7):  return "Q3"
        else:                  return "Q4"

    fy_label = sheet_name.upper()
    long_df["FY"] = long_df["Month"].map(
        lambda m: f"{fy_label} {_quarter(m)}" if pd.notna(m) else None
    )

    # Add GTK Supplier column
    long_df.insert(0, "GTK Supplier", supplier_name)

    # Strip string columns
    for col in long_df.select_dtypes(include="object").columns:
        long_df[col] = long_df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

    # ── Cleanup: drop unwanted columns ───────────────────────────────────────
    _month_names = set(MONTH_MAP.keys())  # jan, feb, … dec

    def _should_drop(col_name: str) -> bool:
        c = str(col_name).strip()
        # Drop blank / unnamed placeholder columns
        if not c or c.startswith("__col"):
            return True
        # Drop columns whose name is a month abbreviation (case-insensitive)
        if c.lower() in _month_names:
            return True
        return False

    cols_to_drop = [c for c in long_df.columns if _should_drop(c)]
    if cols_to_drop:
        long_df.drop(columns=cols_to_drop, inplace=True)

    # Drop columns that are entirely empty (all NaN / blank string)
    empty_cols = [
        c for c in long_df.columns
        if long_df[c].apply(lambda x: x is None or (isinstance(x, float) and pd.isna(x)) or str(x).strip() == "").all()
    ]
    if empty_cols:
        long_df.drop(columns=empty_cols, inplace=True)

    long_df.reset_index(drop=True, inplace=True)
    long_df.drop_duplicates(inplace=True)

    return long_df


def _to_num(val):
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def copy_source_file(src_path: str, source_data_dir: str) -> str:
    """Copy supplier file to source_data dir. Returns dest path."""
    os.makedirs(source_data_dir, exist_ok=True)
    dest = os.path.join(source_data_dir, Path(src_path).name)
    shutil.copy2(src_path, dest)
    return dest


def consolidate_suppliers(
    supplier_files: dict,  # {supplier: file_path}
    sheet_name: str,
    source_data_dir: str,
) -> pd.DataFrame:
    """
    Process all suppliers for a given sheet name, merge into one DataFrame.
    Also copies source files into source_data_dir.
    """
    frames = []
    for supplier, file_path in supplier_files.items():
        copy_source_file(str(file_path), source_data_dir)
        df = process_supplier_sheet(str(file_path), sheet_name, supplier)
        if not df.empty:
            frames.append(df)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)
